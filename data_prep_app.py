r"""
BEI Swing Engine v8.0 — Data Preparation Tool (Standalone)

Standalone Streamlit app for cleaning and merging CSV data from Yahoo Finance.
Does NOT require the full analysis engine. Use this to prepare data before analysis.

Usage:
    python -m streamlit run data_prep_app.py --server.port=8503
"""

import os
import sys
import tempfile
import io

import streamlit as st
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from bei_swing_engine_v8.cleaner import clean_csv_text, rows_to_csv_string, write_cleaned_csv
from bei_swing_engine_v8.merger import merge_csv, parse_cleaned_csv
from bei_swing_engine_v8.logging_config import setup_logging

setup_logging(level="WARNING")

st.set_page_config(
    page_title="Data Preparation — BEI Swing Engine",
    page_icon="🧹",
    layout="wide",
)


def main():
    st.title("🧹 Data Preparation Tool")
    st.markdown("Bersihkan dan gabungkan CSV dari Yahoo Finance. **Standalone** — tidak perlu engine.")
    st.divider()

    mode = st.sidebar.radio("Mode", ["CSV Cleaner", "CSV Merger", "Export XLSX"], index=0)

    if mode == "CSV Cleaner":
        render_cleaner()
    elif mode == "CSV Merger":
        render_merger()
    elif mode == "Export XLSX":
        render_xlsx_export()


def render_cleaner():
    st.header("CSV Cleaner")
    st.markdown("""
    **Langkah:**
    1. Download CSV dari [Yahoo Finance](https://finance.yahoo.com/)
    2. Upload file di bawah ini
    3. Klik **Clean** — CSV otomatis dibersihkan ke format standar
    4. Download hasilnya
    """)

    uploaded = st.file_uploader("Upload CSV mentah (dari Yahoo Finance)", type=["csv"], accept_multiple_files=True)

    if uploaded:
        for f in uploaded:
            st.subheader(f"📄 {f.name}")
            content = f.read().decode("utf-8-sig")
            f.seek(0)

            # Show raw preview
            with st.expander("Lihat data mentah (5 baris pertama)"):
                raw_lines = content.strip().split("\n")[:6]
                st.text("\n".join(raw_lines))

            if st.button(f"Clean {f.name}", key=f"clean_{f.name}"):
                result = clean_csv_text(content, f.name)
                if result.error:
                    st.error(f"Gagal: {result.error}")
                else:
                    st.success(f"Berhasil! {result.row_count} rows, delimiter={result.delimiter}, source={result.source}")
                    cleaned_csv = rows_to_csv_string(result.rows)

                    # Show preview
                    st.dataframe(pd.read_csv(io.StringIO(cleaned_csv)).head(10), use_container_width=True)

                    # Download button
                    output_name = f.name.rsplit(".", 1)[0]
                    if not output_name.endswith("_cleaned"):
                        output_name += "_cleaned"
                    st.download_button(
                        label=f"Download {output_name}.csv",
                        data=cleaned_csv,
                        file_name=f"{output_name}.csv",
                        mime="text/csv",
                    )

                    # XLSX export
                    try:
                        import openpyxl
                        from io import BytesIO
                        df = pd.read_csv(io.StringIO(cleaned_csv))
                        buffer = BytesIO()
                        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                            df.to_excel(writer, sheet_name=output_name[:31], index=False)
                        st.download_button(
                            label=f"Download {output_name}.xlsx (arsip)",
                            data=buffer.getvalue(),
                            file_name=f"{output_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    except ImportError:
                        st.info("Untuk export XLSX: pip install openpyxl")


def render_merger():
    st.header("CSV Merger")
    st.markdown("""
    **Langkah:**
    1. Upload CSV **existing** (sudah dibersihkan)
    2. Upload CSV **baru** (boleh mentah atau sudah clean)
    3. Klik **Merge** — data baru di-append ke existing
    4. Download hasilnya
    """)

    existing_file = st.file_uploader("1. Upload CSV Existing (cleaned)", type=["csv"], key="existing")
    new_files = st.file_uploader("2. Upload CSV Baru", type=["csv"], accept_multiple_files=True, key="new")

    if existing_file and new_files and st.button("Merge"):
        existing_text = existing_file.read().decode("utf-8-sig")
        existing_file.seek(0)

        new_texts = []
        for f in new_files:
            text = f.read().decode("utf-8-sig")
            f.seek(0)
            new_texts.append((f.name, text))

        result = merge_csv(existing_text, new_texts)

        if result.error:
            st.error(f"Gagal: {result.error}")
        else:
            st.success(f"Berhasil! Existing: {result.existing_count}, Baru: {result.new_count}, Total: {result.merged_count}")
            merged_csv = rows_to_csv_string(result.rows)

            # Show preview
            st.dataframe(pd.read_csv(io.StringIO(merged_csv)).tail(10), use_container_width=True)

            # Download
            base_name = existing_file.name.rsplit(".", 1)[0]
            st.download_button(
                label=f"Download {base_name}_merged.csv",
                data=merged_csv,
                file_name=f"{base_name}_merged.csv",
                mime="text/csv",
            )

            # XLSX export
            try:
                import openpyxl
                from io import BytesIO
                df = pd.read_csv(io.StringIO(merged_csv))
                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                    df.to_excel(writer, sheet_name=base_name[:31], index=False)
                st.download_button(
                    label=f"Download {base_name}_merged.xlsx (arsip)",
                    data=buffer.getvalue(),
                    file_name=f"{base_name}_merged.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except ImportError:
                st.info("Untuk export XLSX: pip install openpyxl")


def render_xlsx_export():
    st.header("Export CSV ke XLSX")
    st.markdown("""
    **Langkah:**
    1. Upload satu atau lebih CSV yang sudah clean
    2. Klik **Export** — semua CSV digabung jadi 1 file XLSX (1 sheet per ticker)
    3. Download untuk arsip
    """)

    uploaded = st.file_uploader("Upload CSV files", type=["csv"], accept_multiple_files=True, key="xlsx_export")

    if uploaded and st.button("Export ke XLSX"):
        try:
            import openpyxl
            from io import BytesIO

            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            for f in uploaded:
                content = f.read().decode("utf-8-sig")
                f.seek(0)
                df = pd.read_csv(io.StringIO(content))

                sheet_name = f.name.rsplit(".", 1)[0].replace("_cleaned", "")[:31]
                ws = wb.create_sheet(title=sheet_name)

                for col_idx, col_name in enumerate(df.columns, 1):
                    ws.cell(row=1, column=col_idx, value=col_name)
                for row_idx, row in enumerate(df.itertuples(index=False), 2):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

            buffer = BytesIO()
            wb.save(buffer)

            st.success(f"Berhasil! {len(uploaded)} sheet(s)")
            st.download_button(
                label="Download arsip.xlsx",
                data=buffer.getvalue(),
                file_name="cleaned_archive.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except ImportError:
            st.error("openpyxl belum terinstall. Run: pip install openpyxl")


if __name__ == "__main__":
    main()
