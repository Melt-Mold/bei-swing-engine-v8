#!/usr/bin/env python
"""
Standalone CSV Cleaner CLI app.
Cleans CSV from any source into standard OHLCV format.

Usage:
    python csv_cleaner_app.py [input_files...] -o output_dir
    python csv_cleaner_app.py --help
"""

import argparse
import sys
import os
import glob

# Add parent dir to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from bei_swing_engine_v8.cleaner import clean_csv_file, clean_csv_text, rows_to_csv_string
from bei_swing_engine_v8.logging_config import setup_logging


def export_to_xlsx(output_dir: str, xlsx_path: str):
    """Export all cleaned CSV files in output_dir to a single XLSX workbook (one sheet per ticker)."""
    try:
        import openpyxl
        import pandas as pd
    except ImportError:
        print("ERROR: openpy+pandas required for XLSX export. Run: pip install openpyxl pandas")
        return

    csv_files = sorted(glob.glob(os.path.join(output_dir, "*_cleaned.csv")))
    if not csv_files:
        print("No cleaned CSV files found to export.")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for csv_path in csv_files:
        sheet_name = os.path.basename(csv_path).replace("_cleaned.csv", "")
        # Excel sheet name max 31 chars, no special chars
        sheet_name = sheet_name[:31].replace("/", "_").replace("\\", "_")
        ws = wb.create_sheet(title=sheet_name)

        df = pd.read_csv(csv_path)
        # Write header
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(xlsx_path)
    print(f"XLSX archive saved: {xlsx_path} ({len(csv_files)} sheets)")


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="SwingFlow v8.0 — Universal CSV Cleaner",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python csv_cleaner_app.py BBRI-history.csv -o cleaned/
  python csv_cleaner_app.py *.csv -o cleaned/
  python csv_cleaner_app.py --glob "data/*.csv" -o cleaned/
        """,
    )
    parser.add_argument("inputs", nargs="*", help="Input CSV file(s)")
    parser.add_argument("--glob", dest="glob_pattern", default=None, help="Glob pattern for input files")
    parser.add_argument("-o", "--output-dir", default=".", help="Output directory (default: current dir)")
    parser.add_argument("--xlsx", action="store_true", help="Also export cleaned data to XLSX for archiving")
    parser.add_argument("--log-level", default="INFO", choices=["DEBUG", "INFO", "WARNING", "ERROR"])

    args = parser.parse_args(argv)

    setup_logging(level=args.log_level)

    # Collect input files
    input_files = list(args.inputs)
    if args.glob_pattern:
        input_files.extend(glob.glob(args.glob_pattern))

    if not input_files:
        parser.error("No input files specified. Provide file paths or use --glob.")

    # Validate
    for f in input_files:
        if not os.path.exists(f):
            print(f"ERROR: File not found: {f}", file=sys.stderr)
            sys.exit(1)

    os.makedirs(args.output_dir, exist_ok=True)

    total_files = 0
    total_rows = 0
    errors = 0

    print(f"SwingFlow v8.0 — Universal CSV Cleaner")
    print(f"Processing {len(input_files)} file(s)...")
    print()

    for path in input_files:
        result = clean_csv_file(path, args.output_dir)
        if result.error:
            print(f"  [FAIL] {os.path.basename(path)} — {result.error}")
            errors += 1
        else:
            print(f"  [ OK ] {os.path.basename(path)} → {result.output_name} ({result.row_count} rows)")
            print(f"         delimiter={result.delimiter}, date={result.date_format}, source={result.source}")
            total_files += 1
            total_rows += result.row_count

    print()
    print(f"Summary: {total_files} files cleaned, {total_rows} total rows, {errors} errors")
    if total_files > 0:
        print(f"Output directory: {os.path.abspath(args.output_dir)}")

    # Export to XLSX if requested
    if args.xlsx and total_files > 0:
        xlsx_path = os.path.join(args.output_dir, "cleaned_archive.xlsx")
        export_to_xlsx(args.output_dir, xlsx_path)


if __name__ == "__main__":
    main()
